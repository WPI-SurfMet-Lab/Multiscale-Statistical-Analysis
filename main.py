import warnings
import wx
import wx.adv
import wx.grid
import openpyxl

from Workbook import Workbook
from scipy.optimize import OptimizeWarning
from CurveFit import CurveFit
from Dialogs import RegressionDialog
from Dialogs import GraphSelectDialog
from Dialogs import R2byScaleDialog
from Dialogs import SclbyAreaDialog
from Dialogs import XRValuesDialog
from Dialogs import HHPlotDialog
from PlotData import PlotData
from StatsTestsUI import FtestDialog
from StatsTestsUI import TtestDialog
from StatsTestsUI import ANOVAtestDialog

import faulthandler
faulthandler.enable()

__name__ = 'Multiscale Statisitcal Analysis'
__version__ = '0.1.1'
__license__ = 'MIT'
__author__ = 'Matthew Spofford, Nathaniel Rutkowski'
__author_email__ = 'mespofford@wpi.edu'
__url__ = 'https://github.com/MatthewSpofford/Multiscale-Statistical-Analysis'

global wb_counter
wb_counter = 1

# function for show the curve fit dialog and get regression graphs
def OnRegression(event):

    selectedID = getPlotDataID()
    data = tree_menu.GetItemData(selectedID)

    warnings.simplefilter("error", OptimizeWarning)
    try:

        rsdlg = GraphSelectDialog(frame, data.get_results_scale(), data.get_x_regress(), data.get_regress_sets(),
                                  error_txt, cvf)
        rsdlg.CenterOnScreen()
        resid = rsdlg.ShowModal()

        if resid == wx.ID_OK:

            if rsdlg.prop1check.IsChecked():
                try:
                    gdlg1 = RegressionDialog(frame, "Proportional", data.get_results_scale(), data.get_x_regress(),
                                             data.get_regress_sets(), cvf, selectedID, tree_menu)
                    gdlg1.get_graph().proportional_fit_plot()
                    tree_menu.AppendItem(selectedID, "Proportional Regression", data=gdlg1)
                except (ZeroDivisionError, RuntimeError, Exception, Warning, TypeError, RuntimeWarning, OptimizeWarning) as e:
                    error_txt.AppendText("Proportional: " + str(e) + '\n')

            if rsdlg.lin1check.IsChecked():

                try:
                    gdlg2 = RegressionDialog(frame, "Linear", data.get_results_scale(), data.get_x_regress(),
                                             data.get_regress_sets(), cvf, selectedID, tree_menu)
                    gdlg2.get_graph().linear_fit_plot()
                    tree_menu.AppendItem(selectedID, "Linear Regression", data=gdlg2)
                except (ZeroDivisionError, RuntimeError, Exception, Warning, TypeError, OptimizeWarning, RuntimeWarning, ValueError) as e:
                    error_txt.AppendText("Linear: " + str(e) + '\n')

            if rsdlg.quad1check.IsChecked():
                gdlg3 = RegressionDialog(frame, "Quadratic", data.get_results_scale(), data.get_x_regress(),
                                         data.get_regress_sets(), cvf, selectedID, tree_menu)
                try:
                    gdlg3.get_graph().quadratic_fit_plot()
                    tree_menu.AppendItem(selectedID, "Quadratic Regression", data=gdlg3)
                except (ZeroDivisionError, RuntimeError, Exception, Warning, TypeError, RuntimeWarning, OptimizeWarning) as e:
                    error_txt.AppendText("Quadratic: " + str(e) + '\n')

            if rsdlg.cubic1check.IsChecked():
                gdlg4 = RegressionDialog(frame, "Cubic", data.get_results_scale(), data.get_x_regress(),
                                         data.get_regress_sets(), cvf, selectedID, tree_menu)
                try:
                    gdlg4.get_graph().cubic_fit_plot()
                    tree_menu.AppendItem(selectedID, "Cubic Regression", data=gdlg4)
                except (ZeroDivisionError, RuntimeError, Exception, Warning, TypeError, RuntimeWarning, OptimizeWarning) as e:
                    error_txt.AppendText("Cubic: " + str(e) + '\n')

            if rsdlg.quart1check.IsChecked():
                gdlg5 = RegressionDialog(frame, "Quartic", data.get_results_scale(), data.get_x_regress(),
                                         data.get_regress_sets(), cvf, selectedID, tree_menu)
                try:
                    gdlg5.get_graph().quartic_fit_plot()
                    tree_menu.AppendItem(selectedID, "Quartic Regression", data=gdlg5)
                except (ZeroDivisionError, RuntimeError, Exception, Warning, TypeError, RuntimeWarning, OptimizeWarning) as e:
                    error_txt.AppendText("Quartic: " + str(e) + '\n')

            if rsdlg.quint1check.IsChecked():
                gdlg6 = RegressionDialog(frame, "Quintic", data.get_results_scale(), data.get_x_regress(),
                                         data.get_regress_sets(), cvf, selectedID, tree_menu)
                try:
                    gdlg6.get_graph().quintic_fit_plot()
                    tree_menu.AppendItem(selectedID, "Quintic Regression", data=gdlg6)
                except (ZeroDivisionError, RuntimeError, Exception, Warning, TypeError, RuntimeWarning, OptimizeWarning) as e:
                    error_txt.AppendText("Quintic: " + str(e) + '\n')

            if rsdlg.pow1check.IsChecked():
                gdlg7 = RegressionDialog(frame, "Power", data.get_results_scale(), data.get_x_regress(),
                                         data.get_regress_sets(), cvf, selectedID, tree_menu)
                try:
                    gdlg7.get_graph().power_fit_plot()
                    tree_menu.AppendItem(selectedID, "Power Regression", data=gdlg7)
                except (ZeroDivisionError, RuntimeError, Exception, Warning, TypeError, RuntimeWarning, OptimizeWarning) as e:
                    error_txt.AppendText("Power: " + str(e) + '\n')

            if rsdlg.inverse1check.IsChecked():
                gdlg8 = RegressionDialog(frame, "Inverse", data.get_results_scale(), data.get_x_regress(),
                                         data.get_regress_sets(), cvf, selectedID, tree_menu)
                try:
                    gdlg8.get_graph().inverse_fit_plot()
                    tree_menu.AppendItem(selectedID, "Inverse Regression", data=gdlg8)
                except (ZeroDivisionError, RuntimeError, Exception, Warning, TypeError, RuntimeWarning, OptimizeWarning) as e:
                    error_txt.AppendText("Inverse: " + str(e) + '\n')

            if rsdlg.insq1check.IsChecked():
                gdlg9 = RegressionDialog(frame, "Inverse Square", data.get_results_scale(), data.get_x_regress(),
                                         data.get_regress_sets(), cvf, selectedID, tree_menu)
                try:
                    gdlg9.get_graph().inverse_squared_fit_plot()
                    tree_menu.AppendItem(selectedID, "Inverse Square Regression", data=gdlg9)
                except (ZeroDivisionError, RuntimeError, Exception, Warning, TypeError, RuntimeWarning, OptimizeWarning) as e:
                    error_txt.AppendText("Inverse Square: " + str(e) + '\n')

            if rsdlg.nexp1check.IsChecked():
                gdlg10 = RegressionDialog(frame, "Natural Exponent", data.get_results_scale(), data.get_x_regress(),
                                          data.get_regress_sets(), cvf, selectedID, tree_menu)
                try:
                    gdlg10.get_graph().naturalexp_fit_plot()
                    tree_menu.AppendItem(selectedID, "Natural Exponent Regression", data=gdlg10)
                except (ZeroDivisionError, RuntimeError, Exception, Warning, TypeError, RuntimeWarning, OptimizeWarning) as e:
                    error_txt.AppendText("Natural Exponent: " + str(e) + '\n')

            if rsdlg.Ln1check.IsChecked():
                gdlg11 = RegressionDialog(frame, "Natural Log", data.get_results_scale(), data.get_x_regress(),
                                          data.get_regress_sets(), cvf, selectedID, tree_menu)
                try:
                    gdlg11.get_graph().loge_fit_plot()
                    tree_menu.AppendItem(selectedID, "Natural Log Regression", data=gdlg11)
                except (ZeroDivisionError, RuntimeError, Exception, Warning, TypeError, RuntimeWarning, OptimizeWarning) as e:
                    error_txt.AppendText("Natural Log: " + str(e) + '\n')

            if rsdlg.b10log1check.IsChecked():
                gdlg12 = RegressionDialog(frame, "Base-10 Log", data.get_results_scale(), data.get_x_regress(),
                                          data.get_regress_sets(), cvf, selectedID, tree_menu)
                try:
                    gdlg12.get_graph().log10_fit_plot()
                    tree_menu.AppendItem(selectedID, "Base-10 Log Regression", data=gdlg12)
                except (ZeroDivisionError, RuntimeError, Exception, Warning, TypeError, RuntimeWarning, OptimizeWarning) as e:
                    error_txt.AppendText("Log10: " + str(e) + '\n')

            if rsdlg.invexp1check.IsChecked():
                gdlg13 = RegressionDialog(frame, "Inverse Exponent", data.get_results_scale(), data.get_x_regress(),
                                          data.get_regress_sets(), cvf, selectedID, tree_menu)
                try:
                    gdlg13.get_graph().inverseexp_fit_plot()
                    tree_menu.AppendItem(selectedID, "Inverse Exponent Regression", data=gdlg13)
                except (ZeroDivisionError, RuntimeError, Exception, Warning, TypeError, RuntimeWarning, OptimizeWarning) as e:
                    error_txt.AppendText("Inverse Exponent: " + str(e) + '\n')

            if rsdlg.sin1check.IsChecked():
                gdlg14 = RegressionDialog(frame, "Sine", data.get_results_scale(), data.get_x_regress(),
                                          data.get_regress_sets(), cvf, selectedID, tree_menu)
                try:
                    gdlg14.get_graph().sin_fit_plot()
                    tree_menu.AppendItem(selectedID, "Sine Regression", data=gdlg14)
                except (ZeroDivisionError, RuntimeError, Exception, Warning, TypeError, RuntimeWarning, OptimizeWarning) as e:
                    error_txt.AppendText("Sine: " + str(e) + '\n')

            if rsdlg.cos1check.IsChecked():
                gdlg15 = RegressionDialog(frame, "Cosine", data.get_results_scale(), data.get_x_regress(),
                                          data.get_regress_sets(), cvf, selectedID, tree_menu)
                try:
                    gdlg15.get_graph().cos_fit_plot()
                    tree_menu.AppendItem(selectedID, "Cosine Regression", data=gdlg15)
                except (ZeroDivisionError, RuntimeError, Exception, Warning, TypeError, RuntimeWarning, OptimizeWarning) as e:
                    error_txt.AppendText("Cosine: " + str(e) + '\n')

            if rsdlg.gauss1check.IsChecked():
                gdlg16 = RegressionDialog(frame, "Gaussian", data.get_results_scale(), data.get_x_regress(),
                                          data.get_regress_sets(), cvf, selectedID, tree_menu)
                try:
                    gdlg16.get_graph().gaussian_fit_plot()
                    tree_menu.AppendItem(selectedID, "Gaussian Regression", data=gdlg16)
                except (ZeroDivisionError, RuntimeError, Exception, Warning, TypeError, RuntimeWarning, OptimizeWarning) as e:
                    error_txt.AppendText("Gaussian: " + str(e) + '\n')

            if rsdlg.prop2check.IsChecked():

                gdlg = R2byScaleDialog(frame, "R^2 by Scale for Proportional Regression", data, error_txt, cvf,
                                       tree_menu, selectedID, 0)
                try:
                    gdlg.get_graph().proportional_plot()
                    tree_menu.AppendItem(selectedID, "Proportional R^2 - Scale", data=gdlg)
                except (ZeroDivisionError, RuntimeError, Exception, Warning, TypeError, RuntimeWarning, OptimizeWarning) as e:
                    error_txt.AppendText("Proportional R^2: " + str(e) + '\n')

            if rsdlg.lin2check.IsChecked():

                gdlg = R2byScaleDialog(frame, "R^2 by Scale for Linear Regression", data, error_txt, cvf, tree_menu,
                                       selectedID, 1)
                try:
                    gdlg.get_graph().linear_plot()
                    tree_menu.AppendItem(selectedID, "Linear R^2 - Scale", data=gdlg)
                except (ZeroDivisionError, RuntimeError, Exception, Warning, TypeError, RuntimeWarning, OptimizeWarning) as e:
                    error_txt.AppendText("Linear R^2: " + str(e) + '\n')

            if rsdlg.quad2check.IsChecked():

                gdlg = R2byScaleDialog(frame, "R^2 by Scale for Quadratic Regression", data, error_txt, cvf, tree_menu,
                                       selectedID, 2)
                try:
                    gdlg.get_graph().quadratic_plot()
                    tree_menu.AppendItem(selectedID, "Quadratic R^2 - Scale", data=gdlg)
                except (ZeroDivisionError, RuntimeError, Exception, Warning, TypeError, RuntimeWarning, OptimizeWarning) as e:
                    error_txt.AppendText("Quadratic R^2: " + str(e) + '\n')

            if rsdlg.cubic2check.IsChecked():

                gdlg = R2byScaleDialog(frame, "R^2 by Scale for Cubic Regression", data, error_txt, cvf, tree_menu,
                                       selectedID, 3)
                try:
                    gdlg.get_graph().cubic_plot()
                    tree_menu.AppendItem(selectedID, "Cubic R^2 - Scale", data=gdlg)
                except (ZeroDivisionError, RuntimeError, Exception, Warning, TypeError, RuntimeWarning, OptimizeWarning) as e:
                    error_txt.AppendText("Cubic R^2: " + str(e) + '\n')

            if rsdlg.quart2check.IsChecked():

                gdlg = R2byScaleDialog(frame, "R^2 by Scale for Quartic Regression", data, error_txt, cvf, tree_menu,
                                       selectedID, 4)
                try:
                    gdlg.get_graph().quartic_plot()
                    tree_menu.AppendItem(selectedID, "Quartic R^2 - Scale", data=gdlg)
                except (ZeroDivisionError, RuntimeError, Exception, Warning, TypeError, RuntimeWarning, OptimizeWarning) as e:
                    error_txt.AppendText("Quartic R^2: " + str(e) + '\n')

            if rsdlg.quint2check.IsChecked():

                gdlg = R2byScaleDialog(frame, "R^2 by Scale for Quintic Regression", data, error_txt, cvf, tree_menu,
                                       selectedID, 5)
                try:
                    gdlg.get_graph().quintic_plot()
                    tree_menu.AppendItem(selectedID, "Quintic R^2 - Scale", data=gdlg)
                except (ZeroDivisionError, RuntimeError, Exception, Warning, TypeError, RuntimeWarning, OptimizeWarning) as e:
                    error_txt.AppendText("Quintic R^2: " + str(e) + '\n')

            if rsdlg.pow2check.IsChecked():

                gdlg = R2byScaleDialog(frame, "R^2 by Scale for Power Regression", data, error_txt, cvf, tree_menu,
                                       selectedID, 6)
                try:
                    gdlg.get_graph().power_plot()
                    tree_menu.AppendItem(selectedID, "Power R^2 - Scale", data=gdlg)
                except (ZeroDivisionError, RuntimeError, Exception, Warning, TypeError, RuntimeWarning, OptimizeWarning) as e:
                    error_txt.AppendText("Power R^2: " + str(e) + '\n')

            if rsdlg.inverse2check.IsChecked():

                gdlg = R2byScaleDialog(frame, "R^2 by Scale for Inverse Regression", data, error_txt, cvf, tree_menu,
                                       selectedID, 7)
                try:
                    gdlg.get_graph().inverse_plot()
                    tree_menu.AppendItem(selectedID, "Inverse R^2 - Scale", data=gdlg)
                except (ZeroDivisionError, RuntimeError, Exception, Warning, TypeError, RuntimeWarning, OptimizeWarning) as e:
                    error_txt.AppendText("Inverse R^2: " + str(e) + '\n')

            if rsdlg.insq2check.IsChecked():

                gdlg = R2byScaleDialog(frame, "R^2 by Scale for Inverse Squared Regression", data, error_txt, cvf,
                                       tree_menu, selectedID, 8)
                try:
                    gdlg.get_graph().inverse_squared_plot()
                    tree_menu.AppendItem(selectedID, "Inverse Square R^2 - Scale", data=gdlg)
                except (ZeroDivisionError, RuntimeError, Exception, Warning, TypeError, RuntimeWarning, OptimizeWarning) as e:
                    error_txt.AppendText("Inverse Squared R^2: " + str(e) + '\n')

            if rsdlg.nexp2check.IsChecked():

                gdlg = R2byScaleDialog(frame, "R^2 by Scale for Natural Exponent Regression", data, error_txt, cvf,
                                       tree_menu, selectedID, 9)
                try:
                    gdlg.get_graph().natural_exp_plot()
                    tree_menu.AppendItem(selectedID, "Natural Exponent R^2 - Scale", data=gdlg)
                except (ZeroDivisionError, RuntimeError, Exception, Warning, TypeError, RuntimeWarning, OptimizeWarning) as e:
                    error_txt.AppendText("Natural Exponent R^2: " + str(e) + '\n')

            if rsdlg.Ln2check.IsChecked():

                gdlg = R2byScaleDialog(frame, "R^2 by Scale for Natural Log Regression", data, error_txt, cvf,
                                       tree_menu, selectedID, 10)
                try:
                    gdlg.get_graph().loge_plot()
                    tree_menu.AppendItem(selectedID, "Natural Log R^2 - Scale", data=gdlg)
                except (ZeroDivisionError, RuntimeError, Exception, Warning, TypeError, RuntimeWarning, OptimizeWarning) as e:
                    error_txt.AppendText("Natural Log R^2: " + str(e) + '\n')

            if rsdlg.b10log2check.IsChecked():

                gdlg = R2byScaleDialog(frame, "R^2 by Scale for Log Regression", data, error_txt, cvf, tree_menu,
                                       selectedID, 11)
                try:
                    gdlg.get_graph().log10_plot()
                    tree_menu.AppendItem(selectedID, "Base-10 R^2 - Scale", data=gdlg)
                except (ZeroDivisionError, RuntimeError, Exception, Warning, TypeError, RuntimeWarning, OptimizeWarning) as e:
                    error_txt.AppendText("Log10 R^2: " + str(e) + '\n')

            if rsdlg.invexp2check.IsChecked():

                gdlg = R2byScaleDialog(frame, "R^2 by Scale for Inverse Exponent Regression", data, error_txt, cvf,
                                       tree_menu, selectedID, 12)
                try:
                    gdlg.get_graph().inverse_exp_plot()
                    tree_menu.AppendItem(selectedID, "Inverse Exponent R^2 - Scale", data=gdlg)
                except (ZeroDivisionError, RuntimeError, Exception, Warning, TypeError, RuntimeWarning, OptimizeWarning) as e:
                    error_txt.AppendText("Inverse Exponent R^2: " + str(e) + '\n')

            if rsdlg.sin2check.IsChecked():

                gdlg = R2byScaleDialog(frame, "R^2 by Scale for Sine Regression", data, error_txt, cvf, tree_menu,
                                       selectedID, 13)
                try:
                    gdlg.get_graph().sin_plot()
                    tree_menu.AppendItem(selectedID, "Sine R^2 - Scale", data=gdlg)
                except (ZeroDivisionError, RuntimeError, Exception, Warning, TypeError, RuntimeWarning, OptimizeWarning) as e:
                    error_txt.AppendText("Sine R^2: " + str(e) + '\n')

            if rsdlg.cos2check.IsChecked():

                gdlg = R2byScaleDialog(frame, "R^2 by Scale for Cosine Regression", data, error_txt, cvf, tree_menu,
                                       selectedID, 14)
                try:
                    gdlg.get_graph().cos_plot()
                    tree_menu.AppendItem(selectedID, "Cosine R^2 - Scale", data=gdlg)
                except (ZeroDivisionError, RuntimeError, Exception, Warning, TypeError, RuntimeWarning, OptimizeWarning) as e:
                    error_txt.AppendText("Cosine R^2: " + str(e) + '\n')

            if rsdlg.gauss2check.IsChecked():

                gdlg = R2byScaleDialog(frame, "R^2 by Scale for Gaussian Regression", data, error_txt, cvf, tree_menu,
                                       selectedID, 15)
                try:
                    gdlg.get_graph().gaussian_plot()
                    tree_menu.AppendItem(selectedID, "Gaussian R^2 - Scale", data=gdlg)
                except (ZeroDivisionError, RuntimeError, Exception, Warning, TypeError, RuntimeWarning, OptimizeWarning) as e:
                    error_txt.AppendText("Gaussian R^2: " + str(e) + '\n')
        # Refresh tree menu to show newly created graphs
        tree_menu.Refresh()
    except (ZeroDivisionError, RuntimeError, Warning, TypeError, RuntimeWarning, OptimizeWarning) as e:
        error_txt.AppendText("Graph: " + str(e) + '\n')

# function to get the x-regression values
def OnData(event):
    data = tree_menu.GetItemData(getPlotDataID())

    datadialog = XRValuesDialog(frame, data.get_x_regress())
    datadialog.CenterOnScreen()
    result = datadialog.ShowModal()

    if result == wx.ID_OK:
        datadialog.SaveString()
        data.set_x_regress(datadialog.get_xvals())

# function to show F-test dialog
def OnFtest(event):
    selectedID = getPlotDataID()
    data = tree_menu.GetItemData(selectedID)

    try:
        dlg = FtestDialog(frame, data, error_txt, tree_menu, selectedID)
    except (ZeroDivisionError) as e:
        error_txt.AppendText("F-test: " + str(e) + '\n')

    dlg.CenterOnScreen()
    dlg.ShowModal()
    tree_menu.Refresh()

# current: paired two tails t-test, which one should I use or options...?
# function to show T-test dialog
def OnTtest(event):
    selectedID = getPlotDataID()
    data = tree_menu.GetItemData(selectedID)

    try:
        dlg = TtestDialog(frame, data, error_txt, tree_menu, selectedID)
    except (ZeroDivisionError) as e:
        error_txt.AppendText("T-test: " + str(e) + '\n')

    dlg.CenterOnScreen()
    dlg.ShowModal()
    tree_menu.Refresh()
    # dlg.PooledVarianceTTest()

# function to show the ANOVA test dialog
def OnANOVA(event):
    selectedID = getPlotDataID()
    data = tree_menu.GetItemData(selectedID)

    try:
        dlg = ANOVAtestDialog(frame, data, error_txt, tree_menu, selectedID)
    except (ZeroDivisionError) as e:
        error_txt.AppendText("Anova: " + str(e) + '\n')

    dlg.CenterOnScreen()
    dlg.ShowModal()
    tree_menu.Refresh()

# function to create the scale area plot
def OnAreaPlot(event):
    selectedID = getPlotDataID()
    data = tree_menu.GetItemData(selectedID)

    if data is None :
        error_txt.AppendText("Area-Scale Graph: No data given\n")

    gdlg17 = SclbyAreaDialog(frame, "Scale by Relative Area", data.get_results_scale(),
                             data.get_relative_area(),
                             data.get_legend_txt(), data)
    try:
        gdlg17.get_graph().draw_plot()
        tree_menu.AppendItem(parent=selectedID, text="Relative Area - Scale", data=gdlg17)
        tree_menu.Refresh()
    except (RuntimeError, ZeroDivisionError, Exception, Warning, TypeError, RuntimeWarning, OptimizeWarning) as e:
         error_txt.AppendText("Area-Scale Graph: " + str(e) + '\n')

# function to create the scale complexity plot
def OnComplexityPlot(event):
    selectedID = getPlotDataID()
    data = tree_menu.GetItemData(selectedID)

    if data is None :
        error_txt.AppendText("Area-Scale Graph: No data given\n")

    gdlg18 = SclbyAreaDialog(frame, "Scale by Complexity", data.get_results_scale(), data.get_complexity(),
                             data.get_legend_txt(), data)
    gdlg18.get_graph().get_axes().set_ylabel("Complexity")
    try:
        gdlg18.get_graph().draw_complexity_plot()
        tree_menu.AppendItem(selectedID, "Complexity - Scale", data=gdlg18)
        tree_menu.Refresh()
    except (RuntimeError, ZeroDivisionError, Exception, Warning, TypeError, OptimizeWarning, RuntimeWarning) as e:
        error_txt.AppendText("Complexity-Scale Graph: " + str(e) + '\n')

def OnHHPlot(event):

    gdlg19 = HHPlotDialog(frame, 'Height-Height Plot')
    gdlg19.CenterOnScreen()
    resid = gdlg19.Show()
    tree_menu.Refresh()

# function to get selected graph on left side of main screen
def OnSelection(event):
    selected = tree_menu.GetItemData(tree_menu.GetSelection())

    if isinstance(selected, PlotData):
        selectedWb = selected.get_wb()
        selected.get_error_text().AppendText("Wb: Switching to -- " + selectedWb.get_name() + "\n")
        grid.SetTable(selectedWb)
        grid.AutoSizeColumns()
        grid.Refresh()
    else:
        selected.CenterOnScreen()
        selected.Show()

# function to rename selected graphs/workbooks
def OnRename(event):
    selectedID = tree_menu.GetSelection()
    selected = tree_menu.GetItemData(selectedID)
    newName = event.GetLabel()

    if isinstance(selected, PlotData):
        error_txt.AppendText("Wb: Renaming `" + selected.get_wb().get_name() + "` to `" + newName + "`\n")
        selected.get_wb().set_name(newName)
    else:
        pass

# function to display dialog about the software
def OnAbout(event):
    version = 'v' + __version__
    description = 'An Open-Source, Python-Based application to perform multi-scale \n' \
                  'regression and discrimination analysis using results from Surfract\n' \
                  'and MountainsMap. Developed in collaboration with Christopher A.\n' \
                  'Brown, Ph.D., PE, and the WPI Surface Metrology Lab. Contact\n' \
                  'mespofford@wpi.edu for details. You can support the development of this.\n' \
                  'software by donating below.\n'

    # Reconfigure author strings to use newline seperate and not comma seperation
    developers = "\n".join(__author__.split(", "))

    aboutInfo = wx.Dialog(frame, wx.ID_ANY, 'About ' + __name__ + ' ' + version, size=(480, 400))

    title_text = wx.StaticText(aboutInfo, wx.ID_ANY, label=__name__ + ' ' + version, pos=(60, 20),
                               style=wx.ALIGN_CENTER_HORIZONTAL)
    title_text.SetFont(wx.Font(wx.FontInfo(14)).Bold())
    description_text = wx.StaticText(aboutInfo, wx.ID_ANY, label=description, pos=(45, 50),
                                     style=wx.ALIGN_CENTER_HORIZONTAL)

    github = wx.adv.HyperlinkCtrl(aboutInfo, id=wx.ID_ANY, label='Open-Source Code',
                                  url='https://github.com/MatthewSpofford/Multiscale-Statistical-Analysis',
                                  pos=(180, 150), style=wx.adv.HL_DEFAULT_STYLE)
    donate = wx.adv.HyperlinkCtrl(aboutInfo, id=wx.ID_ANY, label='Support Development',
                                  url='https://paypal.me/nrutkowski1?locale.x=en_US',
                                  pos=(173, 170), style=wx.adv.HL_DEFAULT_STYLE)
    lab_site = wx.adv.HyperlinkCtrl(aboutInfo, id=wx.ID_ANY, label='WPI Surface Metrology Lab',
                                    url='https://www.surfacemetrology.org/',
                                    pos=(159, 190), style=wx.adv.HL_DEFAULT_STYLE)

    line = wx.StaticLine(aboutInfo, id=wx.ID_ANY, pos=(10, 220), size=(450, -1), style=wx.LI_HORIZONTAL)

    d_title = wx.StaticText(aboutInfo, wx.ID_ANY, label='Developers', pos=(190, 240))
    d_title.SetFont(wx.Font(wx.FontInfo(11)).Bold())
    devs = wx.StaticText(aboutInfo, wx.ID_ANY, label=developers, pos=(178, 260), style=wx.ALIGN_CENTER_HORIZONTAL)

    okbtn = wx.Button(aboutInfo, wx.ID_OK, pos=(365, 325))

    aboutInfo.CenterOnScreen()
    aboutInfo.ShowModal()

# function to open the data files
def OnOpen(event):
    frame.EnableCloseButton(False)
    output = False
    try:
        #     # create the open file dialog, if a file exists the user is able to open it will probably change it so
        #     # only the MountainsMap data file format can be opened
        openFileDialog = wx.FileDialog(frame, "Open",  # wildcard="CSV UTF-8 (Comma delimited) (*.csv)|*.csv" ,# \ "
                                       # "TXT (*.txt)|*.txt",
                                       style=wx.FD_OPEN | wx.FD_FILE_MUST_EXIST | wx.FD_MULTIPLE)
        openFileDialog.CenterOnScreen()
        # shows the dialog on screen
        result = openFileDialog.ShowModal()
        # only opens the file if 'open' in dialog is pressed otherwise if 'cancel' in dialog is pressed closes dialog
        if result == wx.ID_OK:
            # gets the file path
            filepath = openFileDialog.GetPaths()
            data = tree_menu.GetItemData(getPlotDataID())

            # opens the file and reads it
            if filepath[0][len(filepath[0]) - 3:len(filepath[0])] == 'csv':
                data.open_file(filepath)
            if filepath[0][len(filepath[0]) - 3:len(filepath[0])] == 'txt':
                data.open_file2(filepath)
            frame.EnableCloseButton(True)
            output = True
        elif result == wx.ID_CANCEL:
            frame.EnableCloseButton(True)
            output = False
    except (Exception) as e:
        raise e
    finally:
        frame.EnableCloseButton(True)
        return output

def OnNewWB(event):

    global wb_counter
    global root

    grid.ClearGrid()
    d = {}
    table = Workbook(d, 'workbook{}'.format(wb_counter), grid.GetNumberRows(), grid.GetNumberCols())
    data = PlotData(error_txt, grid, table)
    grid.SetTable(table)

    item = tree_menu.AppendItem(root, table.name, data=data)
    tree_menu.SelectItem(item)

    error_txt.AppendText('Wb: New Workbook Created -- ' + table.name + '\n')
    wb_counter += 1

def getPlotDataID() :
    global tree_menu
    selectedID = tree_menu.GetSelection()
    selected = tree_menu.GetItemData(selectedID)

    # Check if currently selected node is not plot data
    # If plotdata is not found, go up the tree
    while not isinstance(selected, PlotData) :
        selectedID = tree_menu.GetItemParent(selectedID)
        selected = tree_menu.GetItemData(selectedID)

    return selectedID

def OnSave(event):
    frame.EnableCloseButton(False)
    output = False
    try:
        selectedID = getPlotDataID()
        selectedWorkbook = tree_menu.GetItemData(selectedID).get_wb()

        saveFileDialog = wx.FileDialog(frame, "Save", selectedWorkbook.name, "xlsx (*.xlsx)|*.xlsx", style=wx.FD_SAVE)
        saveFileDialog.CenterOnScreen()
        # shows the dialog on screen when pushes button
        result = saveFileDialog.ShowModal()
        # only saves the file if 'save' in dialog is pressed otherwise if 'cancel' in dialog is pressed closes dialog
        if result == wx.ID_OK:
            # gets the file path
            filepath = saveFileDialog.GetPath()

            cells = selectedWorkbook.get_data().keys()
            values = selectedWorkbook.get_data().values()

            file = openpyxl.Workbook()
            sheet = file.worksheets[0]

            for item in zip(cells, values):

                sheet.cell(row=item[0][0] + 1, column=item[0][1] + 1).value = str(item[1])

            file.save(filepath)
            output = True
        elif result == wx.ID_CANCEL:
            output = False
    except e:
        raise e
    finally:
        frame.EnableCloseButton(True)
        return output

# function to handle window maximization
def onMaxmizeRestore(event):
    global resized
    global frame
    global width, height

    # Shrinks the window at initial resizing
    if frame.IsMaximized() :
        pass
    else:
        if not resized:
            screen = wx.DisplaySize()
            width = screen[0]
            height = screen[1]
            frame.SetSize(0.0, 0.0, width/1.5, height/1.5, sizeFlags=wx.SIZE_AUTO)
            frame.CenterOnScreen()
            resized = True
        else:
            pass

# function to exit the software
def OnExit(event):
    frame.EnableCloseButton(False)
    exitdialog = wx.MessageDialog(frame, "Are you sure you want to quit?", "Exit",
                                  style=wx.YES | wx.NO | wx.ICON_EXCLAMATION)
    exitdialog.SetSize(300, 200)
    exitdialog.CenterOnScreen()
    result = exitdialog.ShowModal()

    if result == wx.ID_YES:
        frame.Destroy()
        return True

    if result == wx.ID_NO:
        frame.EnableCloseButton(True)
        exitdialog.Destroy()
        return False

app = wx.App(redirect=False)
resized = False

# sets the size of the window to be the size of the users computer screen can be set to integers instead
screen = wx.DisplaySize()
width = screen[0]
height = screen[1]

frame = wx.Frame(None, title='Multiscale Analysis')
frame.SetSize(0.0, 0.0, width, height, sizeFlags=wx.SIZE_AUTO)
frame.SetBackgroundColour('#ffffff')
frame.EnableCloseButton(enable=True)
frame.Bind(wx.EVT_CLOSE, OnExit)
frame.Bind(wx.EVT_SIZE, onMaxmizeRestore)

# this is a bunch of GUI stuff
# general curve fit object which has all of the regression curve functions
cvf = CurveFit()
# create the menu bar and populate it
filemenu = wx.Menu()
openitem = wx.MenuItem(parentMenu=filemenu, id=wx.ID_OPEN, text="Open")
openfile = filemenu.Append(openitem)
new_wb = filemenu.Append(wx.ID_ANY, 'New Workbook', 'New Workbook')
frame.Bind(wx.EVT_MENU, OnNewWB, new_wb)
save = wx.MenuItem(filemenu, wx.ID_SAVEAS, 'Save As')
frame.Bind(wx.EVT_MENU, OnSave, save)
filemenu.Append(save)
# bind functions to menu objects
about = wx.MenuItem(filemenu, wx.ID_ABOUT, 'About')
filemenu.Append(about)
exitprogram = wx.MenuItem(filemenu, wx.ID_CLOSE, 'Exit')
filemenu.Append(exitprogram)


# Regression menu initialization
regres_menu = wx.Menu()
xyvals = regres_menu.Append(wx.ID_ANY, 'Regression Values', 'Regression Values')
regression = regres_menu.Append(wx.ID_ANY, 'Curve Fit', 'Curve Fit')
# bind functions to menu objects
frame.Bind(wx.EVT_MENU, OnData, xyvals)
frame.Bind(wx.EVT_MENU, OnRegression, regression)

# Discrimination menu initialization
discrim_menu = wx.Menu()
ftest = discrim_menu.Append(wx.ID_ANY, 'F-test', 'F-test')
ttest = discrim_menu.Append(wx.ID_ANY, 'T-test', 'T-test')
anova = discrim_menu.Append(wx.ID_ANY, 'ANOVA (one-way)', 'ANOVA (one-way)')
# bind functions to menu objects
frame.Bind(wx.EVT_MENU, OnFtest, ftest)
frame.Bind(wx.EVT_MENU, OnTtest, ttest)
frame.Bind(wx.EVT_MENU, OnANOVA, anova)

graphmenu = wx.Menu()
area_scale = graphmenu.Append(wx.ID_ANY, 'Area - Scale Plot', 'Area - Scale Plot')
comp_scale = graphmenu.Append(wx.ID_ANY, 'Complexity - Scale Plot', 'Complexity - Scale Plot')
HHplot = graphmenu.Append(wx.ID_ANY, 'Height-Height Plot', 'Height-Height Plot')
frame.Bind(wx.EVT_MENU, OnAreaPlot, area_scale)
frame.Bind(wx.EVT_MENU, OnComplexityPlot, comp_scale)
frame.Bind(wx.EVT_MENU, OnHHPlot, HHplot)

menuBar = wx.MenuBar()
menuBar.Append(filemenu, 'File')
menuBar.Append(regres_menu, 'Regression')
menuBar.Append(discrim_menu, 'Discrimination')
menuBar.Append(graphmenu, 'Graphs')
frame.SetMenuBar(menuBar)
frame.Bind(wx.EVT_MENU, OnOpen, openfile)
frame.Bind(wx.EVT_MENU, OnExit, exitprogram)
frame.Bind(wx.EVT_CLOSE, OnExit)
frame.Bind(wx.EVT_MENU, OnAbout, about)

# splits the main window into 3 sections, main empty space, side bar with graphs, error logging window
vsplitter = wx.SplitterWindow(frame)
vsplitter.SetBackgroundColour('#ffffff')
hsplitter = wx.SplitterWindow(vsplitter)


# -------------------------------------------------------------------------------------------------------
v_sizer = wx.BoxSizer(wx.VERTICAL)
h_sizer = wx.BoxSizer(wx.VERTICAL)


# main panel with workbook view
main_panel = wx.Panel(hsplitter, style=wx.SIMPLE_BORDER)
main_panel.SetBackgroundColour('#ffffff')

main_sizer = wx.BoxSizer(wx.VERTICAL)

# --------------------------------------------------------------------------------------------------
h_sizer.Add(main_panel, 1, wx.EXPAND)
main_panel.Layout()
main_panel.SetSizerAndFit(main_sizer)

# error panel for error logging
error_panel = wx.Panel(hsplitter, style=wx.BORDER_SUNKEN)

# ------------------------------------------------------------------------------------------------------
h_sizer.Add(error_panel, 1, wx.EXPAND)

# left panel for holding workbook trees on left
left_panel = wx.Panel(vsplitter, style=wx.SIMPLE_BORDER)

# ---------------------------------------------------------------------------------------------------------
v_sizer.Add(left_panel, 1, wx.EXPAND)

# create the error text box which displays the text for errors
error_sizer = wx.BoxSizer(wx.VERTICAL)
error_txt = wx.TextCtrl(error_panel, style=wx.TE_READONLY | wx.TE_MULTILINE)
error_txt.SetBackgroundColour('#000000')
error_txt.SetForegroundColour('#ff8d00')
error_sizer.Add(error_txt, 1, wx.EXPAND)
error_panel.SetSizer(error_sizer)
hsplitter.SplitHorizontally(main_panel, error_panel, sashPosition=580)
vsplitter.SplitVertically(left_panel, hsplitter, sashPosition=200)
sizer = wx.BoxSizer(wx.VERTICAL)
sizer.Add(vsplitter, 1, wx.EXPAND)
# tree which contains the graphs when created, names can be editted
tree_sizer = wx.BoxSizer(wx.VERTICAL)
tree_menu = wx.TreeCtrl(left_panel, style=wx.TR_HAS_BUTTONS | wx.TR_HIDE_ROOT | wx.TR_LINES_AT_ROOT | wx.TR_SINGLE |
                                          wx.TR_FULL_ROW_HIGHLIGHT | wx.TR_EDIT_LABELS)
root = tree_menu.AddRoot("graphs-root")
tree_sizer.Add(tree_menu, 1, wx.EXPAND)
left_panel.SetSizer(tree_sizer)
frame.Bind(wx.EVT_TREE_ITEM_ACTIVATED, OnSelection, tree_menu)
frame.Bind(wx.EVT_TREE_END_LABEL_EDIT, OnRename, tree_menu)

main_sizer.Clear()
main_sizer.Layout()
grid = wx.grid.Grid(main_panel, wx.ID_ANY, style=wx.HSCROLL | wx.VSCROLL | wx.ALWAYS_SHOW_SB)

main_sizer.Add(grid, 1, wx.EXPAND)
main_panel.SetSizer(main_sizer)
grid.CreateGrid(1000, 100)

OnNewWB(wx.EVT_ACTIVATE_APP)

frame.Maximize(True)
frame.CenterOnScreen()
frame.Layout()
frame.SetSizer(sizer)
frame.Show()
app.MainLoop()