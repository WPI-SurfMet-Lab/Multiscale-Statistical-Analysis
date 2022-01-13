import sys
import traceback
import warnings
from enum import Enum

import wx
import wx.adv
import openpyxl
from wx import TreeEvent

import MultiscaleImporter
import MountainsImporter.ImportUtils as ImportUtils

from MultiscaleData import MultiscaleDataset, DatasetAppendOutput, MultiscaleData
from Workbook import Workbook
from scipy.optimize import OptimizeWarning
from Dialogs import RegressionDialog, GraphSelectDialog, R2byScaleDialog, XRValuesDialog, HHPlotDialog
from CanvasPanel import RegressionPlot as RP, SclbyAreaPlot
from CanvasPanel import R2byScalePlot as R2
from StatsTestsUI import FtestDialog, TtestDialog, ANOVAtestDialog

warnings.simplefilter("error", RuntimeWarning)

name = 'Multiscale Statisitcal Analysis'
__version__ = '0.5.0'
__license__ = 'MIT'
__author__ = 'Matthew Spofford, Nathaniel Rutkowski'
__author_email__ = 'mespofford@wpi.edu'
__url__ = 'https://github.com/MatthewSpofford/Multiscale-Statistical-Analysis'

frame = None
app = None

_startup = False

_wb_counter = 1
_wb_set = set()
_main_panel = None

selected_wb = Workbook('')
_selected_wb_ID = None
_selected_results_ID = None
_selected_surfaces_ID = None


class WorkbookIDs:
    def __init__(self, wb_id, surfaces_id, results_id):
        self.wb_id = wb_id
        self.surfaces_id = surfaces_id
        self.results_id = results_id


# Key is Workbook; Value is WorkbookIDs
_workbook_ID_map = {}


# function for show the curve fit dialog and get regression graphs
def OnRegression(event):
    global selected_wb
    dataset = selected_wb.dataset

    warnings.simplefilter("error", OptimizeWarning)
    try:
        rsdlg = GraphSelectDialog(frame, dataset.get_results_scale(), dataset.regress_table,
                                  dataset.get_regress_sets())
        rsdlg.CenterOnScreen()
        resid = rsdlg.ShowModal()

        if resid == wx.ID_OK:
            # Contains list of regression/R^2 dialog properties, checked value functions, and graph generators
            # Tuple Layout
            #   [0] - IsChecked property retrival function
            #   [1] - Dialog title
            #   [2] - Fit plot function to generate regression
            #   [3] - Tree menu item label
            #   [4] - Error dialog label
            #   [5] - Boolean is true if the dialog is for an R^2 dialog
            regressR2Dialogs = (
                (rsdlg.prop1check.IsChecked, "Proportional", RP.proportional_fit_plot, "Proportional Regression",
                 "Proportional", False),
                (rsdlg.lin1check.IsChecked, "Linear", RP.linear_fit_plot, "Linear Regression", "Linear", False),
                (rsdlg.quad1check.IsChecked, "Quadratic", RP.quadratic_fit_plot, "Quadratic Regression", "Quadratic",
                 False),
                (rsdlg.cubic1check.IsChecked, "Cubic", RP.cubic_fit_plot, "Cubic Regression", "Cubic", False),
                (rsdlg.quart1check.IsChecked, "Quartic", RP.quartic_fit_plot, "Quartic Regression", "Quartic", False),
                (rsdlg.quint1check.IsChecked, "Quintic", RP.quintic_fit_plot, "Quintic Regression", "Quintic", False),
                (rsdlg.pow1check.IsChecked, "Power", RP.power_fit_plot, "Power Regression", "Power", False),
                (rsdlg.inverse1check.IsChecked, "Inverse", RP.inverse_fit_plot, "Inverse Regression", "Inverse", False),
                (rsdlg.insq1check.IsChecked, "Inverse Square", RP.inverse_squared_fit_plot, "Inverse Square Regression",
                 "Inverse Square", False),
                (rsdlg.nexp1check.IsChecked, "Natural Exponent", RP.naturalexp_fit_plot, "Natural Exponent Regression",
                 "Natural Exponent", False),
                (rsdlg.Ln1check.IsChecked, "Natural Log", RP.loge_fit_plot, "Natural Log Regression", "Natural Log",
                 False),
                (rsdlg.b10log1check.IsChecked, "Base-10 Log", RP.log10_fit_plot, "Base-10 Log Regression", "Log10",
                 False),
                (rsdlg.invexp1check.IsChecked, "Inverse Exponent", RP.inverseexp_fit_plot,
                    "Inverse Exponent Regression",
                    "Inverse Exponent", False),
                (rsdlg.sin1check.IsChecked, "Sine", RP.sin_fit_plot, "Sine Regression", "Sine", False),
                (rsdlg.cos1check.IsChecked, "Cosine", RP.cos_fit_plot, "Cosine Regression", "Cosine", False),
                (rsdlg.gauss1check.IsChecked, "Gaussian", RP.gaussian_fit_plot, "Gaussian Regression", "Gaussian",
                 False),
                (rsdlg.prop2check.IsChecked, "R^2 by Scale for Proportional Regression", R2.proportional_plot,
                 "Proportional R^2 - Scale", "Proportional R^2", True),
                (rsdlg.lin2check.IsChecked, "R^2 by Scale for Linear Regression", R2.linear_plot, "Linear R^2 - Scale",
                 "Linear R^2", True),
                (rsdlg.quad2check.IsChecked, "R^2 by Scale for Quadratic Regression", R2.quadratic_plot,
                 "Quadratic R^2 - Scale", "Quadratic R^2", True),
                (rsdlg.cubic2check.IsChecked, "R^2 by Scale for Cubic Regression", R2.cubic_plot, "Cubic R^2 - Scale",
                 "Cubic R^2", True),
                (rsdlg.quart2check.IsChecked, "R^2 by Scale for Quartic Regression", R2.quartic_plot,
                 "Quartic R^2 - Scale", "Quartic R^2", True),
                (rsdlg.quint2check.IsChecked, "R^2 by Scale for Quintic Regression", R2.quintic_plot,
                 "Quintic R^2 - Scale", "Quintic R^2", True),
                (rsdlg.pow2check.IsChecked, "R^2 by Scale for Power Regression", R2.power_plot, "Power R^2 - Scale",
                 "Power R^2", True),
                (rsdlg.inverse2check.IsChecked, "R^2 by Scale for Inverse Regression", R2.inverse_plot,
                 "Inverse R^2 - Scale", "Inverse R^2", True),
                (rsdlg.insq2check.IsChecked, "R^2 by Scale for Inverse Squared Regression", R2.inverse_squared_plot,
                 "Inverse Square R^2 - Scale", "Inverse Squared R^2", True),
                (rsdlg.nexp2check.IsChecked, "R^2 by Scale for Natural Exponent Regression", R2.natural_exp_plot,
                 "Natural Exponent R^2 - Scale", "Natural Exponent R^2", True),
                (rsdlg.Ln2check.IsChecked, "R^2 by Scale for Natural Log Regression", R2.loge_plot,
                 "Natural Log R^2 - Scale", "Natural Log R^2", True),
                (rsdlg.b10log2check.IsChecked, "R^2 by Scale for Log Regression", R2.log10_plot, "Base-10 R^2 - Scale",
                 "Log10 R^2", True),
                (rsdlg.invexp2check.IsChecked, "R^2 by Scale for Inverse Exponent Regression", R2.inverse_exp_plot,
                 "Inverse Exponent R^2 - Scale", "Inverse Exponent R^2", True),
                (rsdlg.sin2check.IsChecked, "R^2 by Scale for Sine Regression", R2.sin_plot, "Sine R^2 - Scale",
                 "Sine R^2", True),
                (rsdlg.cos2check.IsChecked, "R^2 by Scale for Cosine Regression", R2.cos_plot, "Cosine R^2 - Scale",
                 "Cosine R^2", True),
                (rsdlg.gauss2check.IsChecked, "R^2 by Scale for Gaussian Regression", R2.gaussian_plot,
                 "Gaussian R^2 - Scale", "Gaussian R^2", True))

            # Run through all given dialogs to find if the given option has been selected
            # If it has been selected, generate the corresponding graph dialog, and graph
            id = 0
            for dialog in regressR2Dialogs:
                checked_func, title, fit_func, menu_label, error_label, isR2 = dialog
                # Continue if this plot fit option been selected
                if checked_func():
                    try:
                        # Generate either R^2 or Regression dialog depending on selection
                        if isR2:
                            gdlg = R2byScaleDialog(frame, title, dataset, tree_menu, _selected_results_ID, id)
                        else:
                            gdlg = RegressionDialog(frame, title, dataset.get_results_scale(), rsdlg.get_x_rvals(),
                                                    dataset.get_regress_sets(), _selected_results_ID, tree_menu)
                        fit_func(gdlg.get_graph())
                        tree_menu.AppendItem(_selected_results_ID, menu_label, data=gdlg)
                        break
                    except (ZeroDivisionError, RuntimeError, Exception, Warning, TypeError, RuntimeWarning,
                            OptimizeWarning) as e:
                        errorMsg(error_label, str(e), str(e))
                        traceback.print_exc()
                # Don't increase ID if current dialog is not an R^2 dialog
                id += 1 if isR2 else 0

            # Refresh tree menu to show newly created graphs
            tree_menu.Refresh()
    except (ZeroDivisionError, RuntimeError, Exception, Warning, TypeError, RuntimeWarning, OptimizeWarning) as e:
        errorMsg("Graph", str(e), str(e))
        if __debug__:
            traceback.print_exc()


# function to get the x-regression values
def OnRegressionValues(event):
    global selected_wb
    dataset = selected_wb.dataset

    regress_val_dialog = XRValuesDialog(frame, dataset.regress_table.copy(), dataset.datasets)
    regress_val_dialog.CenterOnScreen()
    result = regress_val_dialog.ShowModal()

    if result == wx.ID_OK:
        dataset.regress_table = regress_val_dialog.regress_vals_copy


class DiscrimTests:
    """Contains list of discrimination test dialog properties
    [0] - Function for creating the dialogs
    [1] - Error dialog label"""
    Ftest = (FtestDialog, "F-test")
    Ttest = (TtestDialog, "T-test")
    Anova = (ANOVAtestDialog, "Anova")


def OnDiscrimTests(test_choice):
    global selected_wb
    selected_test_func, test_str = test_choice
    dataset = selected_wb.dataset

    try:
        dlg = selected_test_func(frame, dataset, tree_menu, _selected_results_ID, selected_wb)
    except (ZeroDivisionError, RuntimeError, Exception, Warning, TypeError, RuntimeWarning, OptimizeWarning) as e:
        errorMsg(test_str, str(e))
        if __debug__:
            traceback.print_exc()

    dlg.CenterOnScreen()
    dlg.ShowModal()
    tree_menu.Refresh()


# function to show F-test dialog
def OnFtest(event):
    OnDiscrimTests(DiscrimTests.Ftest)


# current: paired two tails t-test, which one should I use or options...?
# function to show T-test dialog
def OnTtest(event):
    OnDiscrimTests(DiscrimTests.Ttest)


# function to show the ANOVA test dialog
def OnANOVA(event):
    OnDiscrimTests(DiscrimTests.Anova)


class ScalePlots(Enum):
    """Contains list of scale plot dialog properties
    [0] - Error dialog label string
    [1] - Dialog title string
    [2] - Function for creating the dialogs
    [3] - Tree menu label string
    [4] - Graph y-axis label"""
    Area = (
        "Area-Scale Graph", "Scale by Relative Area", MultiscaleDataset.get_relative_area, SclbyAreaPlot.draw_plot,
        "Relative Area - Scale", None)
    Complexity = (
        "Complexity-Scale Graph", "Scale by Complexity", MultiscaleDataset.get_complexity,
        SclbyAreaPlot.draw_complexity_plot, "Complexity - Scale", "Complexity")


def display_graph_frame(graph_panel):
    """Remove currently displayed graph if necessary, and display given graph in main panel."""
    if graph_panel is not None:
        _main_panel.DestroyChildren()
        _main_panel.AddChild(graph_panel)


def OnScalePlot(plot_choice: ScalePlots):
    plot_str, title, scale_func, draw_func, menu_text, y_label = plot_choice.value

    global selected_wb
    if selected_wb is None or selected_wb.dataset.is_empty():
        errorMsg(plot_str, "No surfaces given.")
        return

    selected_wb.clear_graph_panel()

    try:
        dataset = selected_wb.dataset
        selected_wb.graph_panel = SclbyAreaPlot(_main_panel, dataset.get_results_scale(), scale_func(dataset), dataset)

        if y_label is not None:
            selected_wb.graph_panel.get_axes().set_ylabel(y_label)

        draw_func(selected_wb.graph_panel)
        selected_wb.graph_panel.Show()
    except Exception as e:
        if selected_wb.graph_panel is not None:
            _main_panel.RemoveChild(selected_wb.graph_panel)
        raise e


# function to create the scale area plot
def OnAreaPlot(event):
    OnScalePlot(ScalePlots.Area)


# function to create the scale complexity plot
def OnComplexityPlot(event):
    OnScalePlot(ScalePlots.Complexity)


def OnHHPlot(event):
    gdlg19 = HHPlotDialog(frame, 'Height-Height Plot')
    gdlg19.CenterOnScreen()
    resid = gdlg19.Show()
    tree_menu.Refresh()


class RenameDialog(wx.Dialog):
    width = 200
    height = 100

    def __init__(self, item_name):
        self._prev_name = item_name

        wx.Dialog.__init__(self, frame, wx.ID_ANY, "Rename - " + item_name,
                           size=(RenameDialog.width, RenameDialog.height))
        self.options_panel = wx.Panel(self, wx.ID_ANY)
        self.sizer = wx.BoxSizer(wx.VERTICAL)

        # Results folder selection button handling
        self.text = wx.TextCtrl(self.options_panel, wx.ID_ANY, item_name)

        # Completion button initialization
        self.ok_btn = wx.Button(self.options_panel, wx.ID_OK, label="Rename")
        self.cancel_btn = wx.Button(self.options_panel, wx.ID_CANCEL, label="Cancel")

        self.completion_sizer = wx.BoxSizer(wx.HORIZONTAL)
        self.completion_sizer.Add(self.cancel_btn, flag=wx.ALIGN_LEFT)
        self.completion_sizer.Add(self.ok_btn, flag=wx.ALIGN_LEFT)

        # Final initialization of dialog's sizer
        self.sizer.AddStretchSpacer()
        self.sizer.Add(self.text, flag=wx.ALIGN_CENTER)
        self.sizer.AddStretchSpacer()
        self.sizer.Add(self.completion_sizer, flag=wx.ALIGN_CENTER)
        self.sizer.AddStretchSpacer()
        self.options_panel.SetSizerAndFit(self.sizer)

        # Center the screen, and set text box to focus
        self.CenterOnScreen()
        self.text.SetFocus()

    def get_new_name(self):
        return self.text.GetValue()


class TreeMenu(wx.Menu):

    def __init__(self, item, item_id: wx.TreeItemId):
        super().__init__()
        self.parent = frame
        self._item = item
        self._item_id = item_id

        self._delete_btn = wx.MenuItem(self, wx.ID_ANY, "Delete")
        self._rename_btn = wx.MenuItem(self, wx.ID_ANY, "Rename")

        self.Append(self._delete_btn)
        self.Append(self._rename_btn)

        frame.Bind(wx.EVT_MENU, self._on_rename, self._rename_btn)
        frame.Bind(wx.EVT_MENU, self._on_delete, self._delete_btn)

    def _on_rename(self, event: wx.MenuEvent):
        dialog = RenameDialog(self._item.name)
        if dialog.ShowModal() == wx.ID_OK:
            new_name = dialog.get_new_name()
            self._item.name = new_name
            # Update tree menu UI with changes
            tree_menu.SetItemText(self._item_id, new_name)

    def _on_delete(self, event: wx.MenuEvent):
        # Get workbook for the selected item
        item_wb_id = tree_menu.GetItemParent(tree_menu.GetItemParent(self._item_id))
        item_wb = tree_menu.GetItemData(item_wb_id)

        # Remove results from workbook
        item_wb.results.remove(self._item)

        # Delete item from tree
        tree_menu.Delete(self._item_id)


class TreeMenuDataset(TreeMenu):
    def __init__(self, item: Workbook, item_id: wx.TreeItemId):
        super().__init__(item, item_id)

        frame.Bind(wx.EVT_MENU, self._on_delete, self._delete_btn)

    def _on_delete(self, event: wx.MenuEvent):
        # Get workbook for the selected item
        item_wb_id = tree_menu.GetItemParent(tree_menu.GetItemParent(self._item_id))
        item_wb = tree_menu.GetItemData(item_wb_id)

        # Back up dataset list, wipe list
        old_dataset_list = item_wb.dataset.datasets.copy()
        item_wb.dataset.clear()

        # Attempt to add new dataset with the item removed
        new_dataset_list = old_dataset_list.copy()
        new_dataset_list.remove(self._item)
        output = item_wb.dataset.append_data(new_dataset_list)
        # If dataset could not be created, restore to old dataset and cancel deletion process
        if output != DatasetAppendOutput.SUCCESS:
            item_wb.dataset.clear()
            item_wb.dataset.append_data(old_dataset_list)
            errorMsg("Error Deleting Surface Data", "Could not delete specified surface, restoring previous data.")
            return

        # Delete item from tree
        tree_menu.Delete(self._item_id)

        # Delete current graph
        item_wb.clear_graph_panel()


class TreeMenuWorkbook(TreeMenu):

    def __init__(self, item: Workbook, item_id: wx.TreeItemId):
        super().__init__(item, item_id)

        frame.Bind(wx.EVT_MENU, self._on_delete, self._delete_btn)

        self._clear_btn = wx.MenuItem(self, wx.ID_ANY, "Empty Results/Surfaces")
        self.Append(self._clear_btn)
        frame.Bind(wx.EVT_MENU, self._on_clear, self._clear_btn)

    def _on_delete(self, event: wx.MenuEvent):
        # Remove selected workbook from set
        _wb_set.remove(self._item)

        # Update UI
        tree_menu.Delete(_workbook_ID_map[self._item].wb_id)
        # Remove graph from pane
        self._item.clear_graph_panel()

        global selected_wb, _selected_wb_ID, _selected_results_ID, _selected_surfaces_ID
        selected_wb = None
        _selected_wb_ID = None
        _selected_results_ID = None
        _selected_surfaces_ID = None

        # Create new workbook if no workbooks left
        if len(_wb_set) <= 0:
            OnNewWB(None)

    def _on_clear(self, event: wx.MenuEvent):
        # Update UI
        self._item.clear_graph_panel()
        tree_menu.DeleteChildren(_selected_surfaces_ID)
        tree_menu.DeleteChildren(_selected_results_ID)

        # Update workbook
        self._item.dataset.clear()
        self._item.results.clear()


def activate_tree_menu(selected_item: wx.TreeItemData, selected_id: wx.TreeItemId, point):
    if isinstance(selected_item, Workbook):
        frame.PopupMenu(TreeMenuWorkbook(selected_item, selected_id), point)
    elif isinstance(selected_item, MultiscaleData):
        frame.PopupMenu(TreeMenuDataset(selected_item, selected_id), point)
    elif isinstance(selected_item, (TtestDialog, FtestDialog, ANOVAtestDialog)):
        frame.PopupMenu(TreeMenu(selected_item, selected_id), point)
    else:
        pass  # Do nothing otherwise


def on_tree_menu_right_click(event):
    selected_id = event.GetItem()
    selected_item = tree_menu.GetItemData(selected_id)
    activate_tree_menu(selected_item, selected_id, event.GetPoint())


# def on_tree_menu_context(event):
#     selected_id = tree_menu.GetSelection()
#     selected_item = tree_menu.GetItemData(selected_id)
#     activate_tree_menu(selected_item, selected_id, tree_menu.Pos)


def OnSelection(event):
    """Handles on click event for sidebar."""
    global selected_wb
    selected_id = tree_menu.GetSelection()
    selected = tree_menu.GetItemData(selected_id)

    if isinstance(selected, Workbook):
        global _selected_wb_ID, _selected_surfaces_ID, _selected_results_ID

        if selected is selected_wb:
            return  # Skip the rest of the function, user clicked on current workbook

        # Hide current graph being displayed
        selected_wb.hide_graph()

        selected_wb = selected
        _selected_wb_ID = selected_id

        ids = _workbook_ID_map[selected_wb]
        _selected_surfaces_ID = ids.surfaces_id
        _selected_results_ID = ids.results_id

        # Potentially show new graph for selected panel
        selected_wb.show_graph()

    elif isinstance(selected, MultiscaleData):
        pass
    elif selected is None:
        pass
    else:
        selected.CenterOnScreen()
        selected.Show()


# function to display dialog about the software
def OnAbout(event):
    version = 'v' + __version__
    description = 'An Open-Source, Python-Based application to perform multi-scale \n' \
                  'regression and discrimination analysis using results from Sfrax\n' \
                  'and MountainsMap. Developed in collaboration with Christopher A.\n' \
                  'Brown, Ph.D., PE, and the WPI Surface Metrology Lab. Contact\n' \
                  'mespofford@wpi.edu for details. You can support the development of this\n' \
                  'software by donating below.\n'

    # Reconfigure author strings to use newline seperate and not comma seperation
    developers = "\n".join(__author__.split(", "))

    aboutInfo = wx.Dialog(frame, wx.ID_ANY, 'About ' + name + ' ' + version, size=(480, 400))

    title_text = wx.StaticText(aboutInfo, wx.ID_ANY, label=name + ' ' + version, pos=(60, 20),
                               style=wx.ALIGN_CENTER_HORIZONTAL)
    title_text.SetFont(wx.Font(wx.FontInfo(14)).Bold())
    description_text = wx.StaticText(aboutInfo, wx.ID_ANY, label=description, pos=(45, 50),
                                     style=wx.ALIGN_CENTER_HORIZONTAL)

    github = wx.adv.HyperlinkCtrl(aboutInfo, id=wx.ID_ANY, label='Open-Source Code',
                                  url='https://github.com/MatthewSpofford/Multiscale-Statistical-Analysis',
                                  pos=(180, 150), style=wx.adv.HL_DEFAULT_STYLE)
    donate = wx.adv.HyperlinkCtrl(aboutInfo, id=wx.ID_ANY, label='Support Development',
                                  url='https://paypal.me/nrutkowski1?locale.x=en_US',
                                  pos=(173, 170), style=wx.adv.HL_DEFAULT_STYLE)
    lab_site = wx.adv.HyperlinkCtrl(aboutInfo, id=wx.ID_ANY, label='WPI Surface Metrology Lab',
                                    url='https://www.surfacemetrology.org/',
                                    pos=(159, 190), style=wx.adv.HL_DEFAULT_STYLE)

    line = wx.StaticLine(aboutInfo, id=wx.ID_ANY, pos=(10, 220), size=(450, -1), style=wx.LI_HORIZONTAL)

    d_title = wx.StaticText(aboutInfo, wx.ID_ANY, label='Developers', pos=(190, 240))
    d_title.SetFont(wx.Font(wx.FontInfo(11)).Bold())
    devs = wx.StaticText(aboutInfo, wx.ID_ANY, label=developers, pos=(178, 260), style=wx.ALIGN_CENTER_HORIZONTAL)

    okbtn = wx.Button(aboutInfo, wx.ID_OK, pos=(365, 325))

    aboutInfo.CenterOnScreen()
    aboutInfo.ShowModal()


def create_open_dialog(choices):
    """
    Generate file open dialog based on file filtering choices and corresponding handling functions
    """
    global selected_wb
    frame.EnableCloseButton(False)

    try:
        # create the open file dialog
        open_file_dialog = wx.FileDialog(frame,
                                         wildcard="|".join([choice[0] for choice in choices]),
                                         style=wx.FD_OPEN | wx.FD_FILE_MUST_EXIST | wx.FD_MULTIPLE)
        open_file_dialog.CenterOnScreen()
        # shows the dialog on screen
        result = open_file_dialog.ShowModal()
        # only opens the file if 'open' in dialog is pressed otherwise if 'cancel' in dialog is pressed closes dialog
        if result == wx.ID_OK:
            # gets the file path
            selection_index = open_file_dialog.GetCurrentlySelectedFilterIndex()
            filepath = open_file_dialog.GetPaths()

            # Handle the opening and reading of the selected files
            datasets = choices[selection_index][1](filepath)

            # Check if no files were opened (or terminated early)
            if datasets is None or not datasets:
                return

            append_output = selected_wb.dataset.append_data(datasets)
            # Handle errors thrown when appending data
            if not append_output:
                output_val = append_output.get_value()
                # Handle scales being ignored
                if append_output == DatasetAppendOutput.SCALES_IGNORED_ERROR:
                    list_str = [data.name for data in output_val]
                    errorMsg("Dataset", "Scales were ignored when adding " + ", ".join(list_str) + '\n')
                # Handle general error
                else:
                    error_str = "Issue with adding to dataset "
                    if output_val:
                        error_str += output_val
                    errorMsg("Dataset Error", error_str)

        elif result == wx.ID_CANCEL:
            pass
    except Exception as e:
        errorMsg("File Open Error", str(e), str(e))
        if __debug__:
            traceback.print_exc()
    finally:
        frame.EnableCloseButton(True)

    # Remove currently results in tree, and add updated datasets to list
    tree_menu.DeleteChildren(_selected_surfaces_ID)
    for data in selected_wb.dataset.datasets:
        tree_menu.AppendItem(_selected_surfaces_ID, data.name, data=data)
    tree_menu.Expand(_selected_surfaces_ID)


def on_open_surf(event):
    """Function to open the surface files."""
    choices = [("MountainsMap Surface Files (*.sur)|*.sur", MultiscaleImporter.open_sur)]
    create_open_dialog(choices)


def on_open_results(event):
    """Function to open the results files."""
    choices = [("MountainsMap Results Text Files (*.txt)|*.txt", MultiscaleImporter.open_results_file),
               ("Sfrax CSV Results - UTF-8 (*.csv)|*.csv", MultiscaleImporter.open_sfrax)]
    create_open_dialog(choices)


_wkbk_tree_results = "Results"
_wkbk_tree_surfaces = "Surfaces"


def OnNewWB(event):
    global _wb_counter, root, selected_wb, _startup, _selected_wb_ID, _selected_surfaces_ID, _selected_results_ID

    # Don't attempt to hide the graph on startup, because it is not initialized yet
    if not _startup:
        selected_wb.hide_graph()

    selected_wb = Workbook('workbook{}'.format(_wb_counter))
    _selected_wb_ID = tree_menu.AppendItem(root, selected_wb.name, data=selected_wb)
    _selected_surfaces_ID = tree_menu.AppendItem(_selected_wb_ID, _wkbk_tree_surfaces, data=None)
    _selected_results_ID = tree_menu.AppendItem(_selected_wb_ID, _wkbk_tree_results, data=None)

    _workbook_ID_map[selected_wb] = WorkbookIDs(_selected_wb_ID, _selected_surfaces_ID, _selected_results_ID)

    tree_menu.SelectItem(_selected_wb_ID)
    tree_menu.Expand(_selected_wb_ID)
    _wb_counter += 1

    _wb_set.add(selected_wb)


def OnSave(event):
    global selected_wb
    frame.EnableCloseButton(False)
    output = False
    try:
        save_file_dialog = wx.FileDialog(frame, "Save", selected_wb.name, "xlsx (*.xlsx)|*.xlsx",
                                         style=wx.FD_SAVE)
        save_file_dialog.CenterOnScreen()
        # shows the dialog on screen when pushes button
        result = save_file_dialog.ShowModal()
        # only saves the file if 'save' in dialog is pressed otherwise if 'cancel' in dialog is pressed closes dialog
        if result == wx.ID_OK:
            # gets the file path
            filepath = save_file_dialog.GetPath()

            cells = selected_wb.get_table_data().keys()
            values = selected_wb.get_table_data().values()

            file = openpyxl.Workbook()
            sheet = file.worksheets[0]

            for item in zip(cells, values):
                sheet.cell(row=item[0][0] + 1, column=item[0][1] + 1).value = str(item[1])

            file.save(filepath)
            output = True
        elif result == wx.ID_CANCEL:
            output = False
    except e:
        errorMsg("File Save Error", str(e))
    finally:
        frame.EnableCloseButton(True)
        return output


# function to handle window maximization
def OnMaxmizeRestore(event):
    global resized
    global frame
    global width, height

    # Shrinks the window at initial resizing
    if frame.IsMaximized():
        pass
    else:
        if not resized:
            screen = wx.DisplaySize()
            width = screen[0]
            height = screen[1]
            frame.SetSize(0, 0, int(width / 1.5), int(height / 1.5), sizeFlags=wx.SIZE_AUTO)
            frame.CenterOnScreen()
            resized = True
        else:
            pass


# function to exit the software
def OnExit(event):
    frame.EnableCloseButton(False)
    exitdialog = wx.MessageDialog(frame, "Are you sure you want to quit?", "Exit",
                                  style=wx.YES | wx.NO | wx.ICON_EXCLAMATION)
    exitdialog.SetSize(300, 200)
    exitdialog.CenterOnScreen()
    result = exitdialog.ShowModal()

    if result == wx.ID_YES:
        frame.Destroy()
        return True

    if result == wx.ID_NO:
        frame.EnableCloseButton(True)
        exitdialog.Destroy()
        return False

    app.Destroy()


def errorMsg(title, msg, trace_str=None):
    """
    Displays error message dialog. Waits for the message dialog to close, prevents accessing parent frame.
    """
    # Print to stderr the stacktrace
    if trace_str is not None:
        print(trace_str, file=sys.stderr)
    # Create error dialog
    dialog = wx.MessageDialog(frame, str(msg), title, style=wx.ICON_ERROR | wx.OK)
    dialog.ShowModal()


# Displays error message dialog
# Waits for the message dialog to close, prevents accesing parent frame
def warnMsg(title, msg):
    dialog = wx.MessageDialog(frame, msg, title, style=wx.ICON_WARNING | wx.OK)
    dialog.ShowModal()


def exceptionHandler(etype, value, trace):
    trace = traceback.format_exception(etype, value, trace)
    trace_str = "".join(trace)

    errorMsg(etype.__name__ + " Error", value, trace_str)


if __name__ == "__main__":
    app = wx.App(redirect=False)
    resized = False

    # sets the size of the window to be the size of the users computer screen can be set to integers instead
    screen = wx.DisplaySize()
    width = screen[0]
    height = screen[1]

    frame = wx.Frame(None, title='Multiscale Statistical Analysis')
    frame.SetSize(0, 0, width, height, sizeFlags=wx.SIZE_AUTO)
    frame.SetBackgroundColour('#ffffff')
    frame.EnableCloseButton(enable=True)
    frame.Bind(wx.EVT_CLOSE, OnExit)
    frame.Bind(wx.EVT_SIZE, OnMaxmizeRestore)

    # this is a bunch of GUI stuff
    # create the menu bar and populate it
    filemenu = wx.Menu()
    if not isinstance(ImportUtils.find_mountains_map(), ImportUtils.MountainsNotFound):
        open_surf = wx.MenuItem(parentMenu=filemenu, id=wx.ID_ANY, text="Open Surfaces")
        open_surf = filemenu.Append(open_surf)
        frame.Bind(wx.EVT_MENU, on_open_surf, open_surf)
    open_results = wx.MenuItem(parentMenu=filemenu, id=wx.ID_ANY, text="Open Results Files")
    open_results = filemenu.Append(open_results)
    new_wb = filemenu.Append(wx.ID_ANY, 'New Workbook', 'New Workbook')
    frame.Bind(wx.EVT_MENU, OnNewWB, new_wb)

    # save = wx.MenuItem(filemenu, wx.ID_SAVEAS, 'Save As')
    # frame.Bind(wx.EVT_MENU, OnSave, save)
    # filemenu.Append(save)

    # bind functions to menu objects
    about = wx.MenuItem(filemenu, wx.ID_ABOUT, 'About')
    filemenu.Append(about)
    exitprogram = wx.MenuItem(filemenu, wx.ID_CLOSE, 'Exit')
    filemenu.Append(exitprogram)

    # Regression menu initialization
    regres_menu = wx.Menu()
    xyvals = regres_menu.Append(wx.ID_ANY, 'Regression Values', 'Regression Values')
    regression = regres_menu.Append(wx.ID_ANY, 'Curve Fit', 'Curve Fit')
    # bind functions to menu objects
    frame.Bind(wx.EVT_MENU, OnRegressionValues, xyvals)
    frame.Bind(wx.EVT_MENU, OnRegression, regression)

    # Discrimination menu initialization
    discrim_menu = wx.Menu()
    ftest = discrim_menu.Append(wx.ID_ANY, 'F-test', 'F-test')
    ttest = discrim_menu.Append(wx.ID_ANY, 'T-test', 'T-test')
    anova = discrim_menu.Append(wx.ID_ANY, 'ANOVA (one-way)', 'ANOVA (one-way)')
    # bind functions to menu objects
    frame.Bind(wx.EVT_MENU, OnFtest, ftest)
    frame.Bind(wx.EVT_MENU, OnTtest, ttest)
    frame.Bind(wx.EVT_MENU, OnANOVA, anova)

    graph_menu = wx.Menu()
    area_scale = graph_menu.Append(wx.ID_ANY, 'Area - Scale Plot', 'Area - Scale Plot')
    frame.Bind(wx.EVT_MENU, OnAreaPlot, area_scale)

    comp_scale = graph_menu.Append(wx.ID_ANY, 'Complexity - Scale Plot', 'Complexity - Scale Plot')
    frame.Bind(wx.EVT_MENU, OnComplexityPlot, comp_scale)

    # HHplot = graphmenu.Append(wx.ID_ANY, 'Height-Height Plot', 'Height-Height Plot')
    # frame.Bind(wx.EVT_MENU, OnHHPlot, HHplot)

    menuBar = wx.MenuBar()
    menuBar.Append(filemenu, 'File')
    menuBar.Append(regres_menu, 'Regression')
    menuBar.Append(discrim_menu, 'Discrimination')
    menuBar.Append(graph_menu, 'Graphs')
    frame.SetMenuBar(menuBar)
    frame.Bind(wx.EVT_MENU, on_open_results, open_results)
    frame.Bind(wx.EVT_MENU, OnExit, exitprogram)
    frame.Bind(wx.EVT_CLOSE, OnExit)
    frame.Bind(wx.EVT_MENU, OnAbout, about)

    # splits the main window into 2 sections: side bar with , side bar with datasets, and graph window
    vsplitter = wx.SplitterWindow(frame)
    vsplitter.SetBackgroundColour('#DCDCDC')

    v_sizer = wx.BoxSizer(wx.VERTICAL)
    h_sizer = wx.BoxSizer(wx.VERTICAL)

    # main panel with workbook view
    _main_panel = wx.Panel(vsplitter, style=wx.SIMPLE_BORDER)
    _main_panel.Layout()
    main_sizer = wx.BoxSizer(wx.VERTICAL)
    _main_panel.SetSizerAndFit(main_sizer)

    # --------------------------------------------------------------------------------------------------
    h_sizer.Add(_main_panel, 1, wx.EXPAND)

    # ------------------------------------------------------------------------------------------------------

    # left panel for holding workbook trees on left
    left_panel = wx.Panel(vsplitter, style=wx.SIMPLE_BORDER)

    # ---------------------------------------------------------------------------------------------------------
    v_sizer.Add(left_panel, 1, wx.EXPAND)

    # create the error text box which displays the text for errors
    vsplitter.SplitVertically(left_panel, _main_panel, sashPosition=200)
    sizer = wx.BoxSizer(wx.VERTICAL)
    sizer.Add(vsplitter, 1, wx.EXPAND)
    # tree which contains the graphs when created, names can be editted
    tree_sizer = wx.BoxSizer(wx.VERTICAL)
    tree_menu = wx.TreeCtrl(left_panel, style=wx.TR_HAS_BUTTONS | wx.TR_HIDE_ROOT | wx.TR_LINES_AT_ROOT | wx.TR_SINGLE |
                                              wx.TR_FULL_ROW_HIGHLIGHT)
    root = tree_menu.AddRoot("graphs-root")
    tree_sizer.Add(tree_menu, 1, wx.EXPAND)
    left_panel.SetSizer(tree_sizer)
    frame.Bind(wx.EVT_TREE_ITEM_ACTIVATED, OnSelection, tree_menu)
    frame.Bind(wx.EVT_TREE_ITEM_RIGHT_CLICK, on_tree_menu_right_click, tree_menu)
    # frame.Bind(wx.EVT_CONTEXT_MENU, on_tree_menu_context, tree_menu)

    main_sizer.Clear()
    main_sizer.Layout()

    _main_panel.SetSizer(main_sizer)

    OnNewWB(wx.EVT_ACTIVATE_APP)

    frame.Maximize(True)
    frame.CenterOnScreen()
    frame.Layout()
    frame.SetSizer(sizer)
    frame.Show()

    # Assign exception handler
    sys.excepthook = exceptionHandler

    # Check the existence of MountainsMap
    try:
        ImportUtils.find_mountains_map()
    except (FileNotFoundError, Exception) as e:
        warnMsg("MountainsMap Warning", "MountainsMap installation could not be found.")

    # Run Multiscale Analysis App
    app.MainLoop()

else:
    pass
